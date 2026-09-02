#!/usr/bin/env python3
"""
Build annotation spreadsheets from the Khmer polarity corpus.

Source: github.com/ye-kyaw-thu/kh-polarity  (kh-polar.ver1.0.txt)
Format of each line:   sentence ||| polarity-cue ||| label

Six annotators working in three pairs. Each PAIR gets one file of 200 sentences;
both members annotate that same file independently, then reconcile. Different
pairs get different sentences, so 600 distinct sentences are covered in total.

    pair1_annotation.xlsx   200 sentences  ->  annotator A + annotator B
    pair2_annotation.xlsx   200 sentences  ->  annotator C + annotator D
    pair3_annotation.xlsx   200 sentences  ->  annotator E + annotator F

What is removed from the annotator files, and why:

  * the ORIGINAL LABEL - if annotators can see the answer, the agreement score
    measures nothing
  * the POLARITY CUE - the corpus marks which phrase carries the sentiment.
    Showing it hands over most of the judgement.

Both are kept in annotation_KEY.xlsx, which stays with the coordinator until
everybody has submitted.

Sampling: balanced across the three classes. The corpus is 58% positive, and an
annotator who notices that starts guessing positive. Balanced also gives kappa a
fair chance - agreement on a skewed set is flattered by the base rate.

Usage:
    python scripts/build_annotation_sheets.py
    python scripts/build_annotation_sheets.py --per-file 250 --pairs 3
"""
import argparse
import os
import textwrap

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data", "raw", "kh-polar.txt")
OUT = os.path.join(ROOT, "data", "annotation")

LABELS = ["positive", "negative", "neutral"]
SEED = 42

# Khmer needs a font that actually has the glyphs; without this Excel picks a
# fallback and the script renders as boxes or with broken diacritic stacking.
KHMER_FONT = "Noto Sans Khmer"
UI_FONT = "Calibri"

SENT_WIDTH = 78          # visible characters per line in the sentence column
MIN_CHARS, MAX_CHARS = 25, 300

HEADER_FILL = PatternFill("solid", fgColor="2F4858")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name=UI_FONT)
TODO_FILL = PatternFill("solid", fgColor="FFF3CD")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INSTRUCTIONS = [
    ("របៀបដាក់ស្លាក  /  How to annotate", True),
    ("", False),
    ("1. ចូលទៅផ្ទាំង 'Annotation' / Go to the 'Annotation' tab.", False),
    ("2. អានប្រយោគនីមួយៗ / Read each sentence.", False),
    ("3. ជ្រើសរើសស្លាកនៅជួរ LABEL / Pick a label from the dropdown:", False),
    ("", False),
    ("      positive   វិជ្ជមាន  - the sentence expresses a favourable feeling", False),
    ("      negative   អវិជ្ជមាន - the sentence expresses an unfavourable feeling", False),
    ("      neutral    អព្យាក្រឹត - a plain statement, no clear feeling either way", False),
    ("", False),
    ("Rows you have not done stay YELLOW. Label all of them.", False),
    ("", False),
    ("Rules", True),
    ("", False),
    ("  * Judge the SENTENCE AS A WHOLE, not one word inside it.", False),
    ("  * Work ALONE. Do not discuss individual sentences with your partner", False),
    ("    until you have BOTH finished - independent judgements are the point.", False),
    ("  * Factual news reporting is usually neutral, even about a bad event.", False),
    ("    Ask: is the writer expressing a feeling, or just reporting?", False),
    ("  * A sentence can describe something negative in neutral language.", False),
    ("    Label the writer's stance, not the topic.", False),
    ("  * If you genuinely cannot decide, choose neutral.", False),
    ("  * Do NOT save this file as .csv - it will corrupt the Khmer text.", False),
    ("    Keep it as .xlsx.", False),
    ("", False),
    ("When finished", True),
    ("", False),
    ("  Rename the file with your own name, e.g. sokha.xlsx, and send it back.", False),
    ("  Then compare with your partner and agree a final label for each row.", False),
]


def load_corpus() -> pd.DataFrame:
    rows = []
    with open(RAW, encoding="utf-8") as f:
        for line in f:
            parts = [p.strip() for p in line.rstrip("\n").split("|||")]
            if len(parts) == 3 and parts[0] and parts[2] in LABELS:
                rows.append({"sentence": parts[0], "cue": parts[1], "label": parts[2]})
    df = pd.DataFrame(rows)
    df["chars"] = df["sentence"].str.len()
    return df


def estimate_height(text: str, width: int = SENT_WIDTH) -> float:
    lines = max(1, len(textwrap.wrap(str(text), width)))
    # Khmer stacks diacritics above and below, so rows need more room per line
    return min(max(lines * 21 + 10, 34), 409)


def write_sheet(df: pd.DataFrame, path: str, title: str) -> None:
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Instructions"
    ws0.column_dimensions["A"].width = 92
    for i, (line, is_head) in enumerate(INSTRUCTIONS, start=1):
        c = ws0.cell(row=i, column=1, value=line)
        c.font = (Font(bold=True, size=13, color="2F4858", name=KHMER_FONT)
                  if is_head else Font(size=11, name=KHMER_FONT))

    ws = wb.create_sheet("Annotation")
    headers = ["id", "ប្រយោគ  /  SENTENCE", "LABEL", "notes (optional)"]
    widths = [7, SENT_WIDTH, 14, 26]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    khmer = Font(name=KHMER_FONT, size=12)
    wrap = Alignment(wrap_text=True, vertical="top")
    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(row=r, column=1, value=int(row.id)).alignment = Alignment(
            horizontal="center", vertical="top")
        c = ws.cell(row=r, column=2, value=str(row.sentence))
        c.alignment, c.font = wrap, khmer
        ws.cell(row=r, column=3).alignment = Alignment(
            horizontal="center", vertical="center")
        ws.cell(row=r, column=4).alignment = wrap
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        ws.row_dimensions[r].height = estimate_height(row.sentence)

    last = len(df) + 1
    dv = DataValidation(
        type="list", formula1=f'"{",".join(LABELS)}"', allow_blank=True,
        showDropDown=False, showErrorMessage=True,
        errorTitle="Choose from the list",
        error="Pick positive, negative or neutral from the dropdown.",
    )
    ws.add_data_validation(dv)
    dv.add(f"C2:C{last}")
    ws.conditional_formatting.add(
        f"A2:D{last}", FormulaRule(formula=['$C2=""'], fill=TODO_FILL, stopIfTrue=False))
    ws.auto_filter.ref = f"A1:D{last}"

    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--per-file", type=int, default=200)
    ap.add_argument("--pairs", type=int, default=3)
    args = ap.parse_args()

    os.makedirs(OUT, exist_ok=True)
    df = load_corpus()
    print(f"corpus loaded      : {len(df)} sentences")
    print(f"  labels           : {df.label.value_counts().to_dict()}")

    pool = df[df.chars.between(MIN_CHARS, MAX_CHARS)]
    pool = pool[~pool.sentence.duplicated()]
    print(f"  after dedupe + length filter ({MIN_CHARS}-{MAX_CHARS} chars): {len(pool)}")

    total = args.per_file * args.pairs
    per_class = total // len(LABELS)
    parts = []
    for label in LABELS:
        avail = pool[pool.label == label]
        if len(avail) < per_class:
            raise SystemExit(
                f"Only {len(avail)} '{label}' sentences available, need {per_class}. "
                f"Lower --per-file."
            )
        parts.append(avail.sample(per_class, random_state=SEED))
    sample = pd.concat(parts)
    # top up if total is not divisible by 3
    if len(sample) < total:
        extra = pool[~pool.index.isin(sample.index)].sample(
            total - len(sample), random_state=SEED)
        sample = pd.concat([sample, extra])

    sample = sample.sample(frac=1, random_state=SEED).reset_index(drop=True)
    sample.insert(0, "id", range(1, len(sample) + 1))

    print(f"\nsampled            : {len(sample)} sentences, balanced")
    print(f"  {sample.label.value_counts().to_dict()}")

    for p in range(args.pairs):
        chunk = sample.iloc[p * args.per_file:(p + 1) * args.per_file]
        path = os.path.join(OUT, f"pair{p + 1}_annotation.xlsx")
        write_sheet(chunk, path, f"Pair {p + 1}")
        print(f"\n  pair{p + 1}: ids {int(chunk.id.min())}-{int(chunk.id.max())}  "
              f"{len(chunk)} rows  {chunk.label.value_counts().to_dict()}")
        print(f"         {path}")

    key = os.path.join(OUT, "annotation_KEY.xlsx")
    sample["pair"] = [f"pair{i // args.per_file + 1}" for i in range(len(sample))]
    sample[["id", "pair", "label", "cue", "sentence"]].to_excel(key, index=False)
    print(f"\n  KEY (original labels + cues): {key}")
    print("  Keep the KEY back until everyone has submitted.")

    # leak check
    for p in range(args.pairs):
        chk = pd.read_excel(os.path.join(OUT, f"pair{p + 1}_annotation.xlsx"),
                            sheet_name="Annotation")
        leaked = [c for c in chk.columns if str(c).lower() in ("label", "cue") and
                  chk[c].notna().any()]
        assert not leaked, f"pair{p+1} leaked {leaked}"
    print("\nleak check: ok, no original labels or cues in the annotator files")


if __name__ == "__main__":
    main()
